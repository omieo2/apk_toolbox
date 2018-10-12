#!/usr/bin/python3

import getopt
import hashlib
import os
import re
import sys
import time
import requests
import openpyxl
import datetime

from androguard.core.bytecodes.apk import APK
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

now = str(datetime.date.today())
download_dir = os.path.join(os.getcwd(), now)


def main(argv):
    global opts
    try:
        opts, args = getopt.getopt(argv[1:], 'hvd:r:i:', ['help', 'version', 'download=', 'rename=', 'info='])
    except getopt.GetoptError as err:
        print('读取参数时发生错误！', err)

    for name, value in opts:
        if name in ('-h', '--help'):
            usage()
            return
        elif name in ('-v', '--version'):
            version()
            return
        elif name in ('-d', '--download'):
            download(value)
        elif name in ('-r', '--rename'):
            rename2md5(value)
        elif name in ('-i', '--info'):
            app_info(value)


def usage():
    print('———————— 使用说明 ————————')
    print('-h, --help:         帮助信息')
    print('-v, --version:      版本号')
    print('-d, --download:     批量爬虫apk')
    print('-r, --rename:       将文件以md5重命名')
    print('-i, --info:         获取apk信息')


def version():
    print('———————— 版本信息 ————————')
    print('apk工具箱 v1.0.1 build on 2018/10/12.')


def download(filename):
    if not os.path.isfile(filename):
        print('输入参数有误，不是文件！')
        return

    if not os.path.exists(download_dir):
        os.mkdir(download_dir)

    print('Index', '\tResult', '\t\t Url')
    start = time.time()
    count_succeed = 0
    count_failure = 0
    max_row, urls = get_all_links(filename)
    wd = WriteData(filename='download_apk_',
                   navigation_bar=['url', 'result', 'filename', 'file_md5', 'app_name', 'pkg_name', 'cert_md5',
                                   'app_version'])
    wd.write_data(is_init=True)
    for i in range(max_row):
        if download_apk(urls[i], i):
            print(max_row - i, '\tSucceed\t\t', urls[i])
            apk_name = rename_apk(i)
            count_succeed += 1

            excel_data = [urls[i], 'Succeed'] + get_apk_info(apk_name)
            wd.write_data(row=i, data=excel_data)
        else:
            print(max_row - i, '\tFailure\t\t', urls[i])
            count_failure += 1

            excel_data = [urls[i], 'Failure']
            wd.write_data(row=i, data=excel_data)

    print('\nTotal info：\n共%d条链接，成功：%d，失败：%d\t耗时%.1fs' % (max_row, count_succeed, count_failure, time.time() - start))


def get_all_links(filename):
    links = []
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet1']
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        url = ws.cell(row=row, column=1).value
        if not re.match(r'^https?:/{2}\w.+$', url):
            url = "http://" + url
        links.append(url)
    return max_row, links


def download_apk(url, i):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Linux; U; Android 4.0.3; zh-cn; M032 Build/IML74K) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30',
        'Connection': 'keep-alive'}
    try:
        response = requests.get(url, headers=headers, timeout=15)
        # media_type = response.headers['Content-Type']
        # apk_type == 'application/vnd.android.package-archive'

        # 只下载小于 100M 的应用
        file_size = round(float(response.headers['Content-Length']) / 1048576.0, 2)
        if response.status_code == 200 and file_size <= 100:
            tmp_apk_path = os.path.join(download_dir, str(i) + '.apk')
            with open(tmp_apk_path, 'wb') as f:
                f.write(response.content)
            return True
        else:
            return False
    except Exception as exp:
        return False


def rename_apk(i):
    index_filename = str(i) + '.apk'
    apk_path = os.path.join(download_dir, index_filename)
    apk_md5 = get_file_md5(apk_path)
    apk_file = apk_md5 + '.apk'
    new_name = os.path.join(download_dir, apk_file)

    if not os.path.exists(new_name):
        os.rename(apk_path, new_name)
    else:
        os.remove(apk_path)

    return apk_file


def get_file_md5(file_path):
    with open(file_path, 'rb') as f:
        md5obj = hashlib.md5()
        md5obj.update(f.read())
        md5 = md5obj.hexdigest()
        md5 = str(md5).lower()
    return md5


def app_info(path):
    start = time.time()
    if not os.path.exists(download_dir):
        os.mkdir(download_dir)

    if os.path.isdir(path):
        wd = WriteData(filename='apk_info_',
                       navigation_bar=['filename', 'file_md5', 'app_name', 'pkg_name', 'cert_md5', 'app_version'])
        wd.write_data(is_init=True)
        for root, dirs, files in os.walk(path):
            for f in range(len(files)):
                info = get_apk_info(files[f], root)
                if info:
                    print(str(info))
                    wd.write_data(data=info, row=f)

            print('\nTotal info:\n已获取%d个文件信息\t耗时%.2fs' % (len(files), time.time() - start))
    else:
        print('参数输入有误，不是一个目录...')


def get_apk_info(f, root=None):
    """
    获取apk信息
    :param root:
    :param f:
    :return:
    """
    if root:
        apk_path = os.path.join(root, f)
    else:
        apk_path = os.path.join(download_dir, f)

    apk_info = []
    try:
        apk = APK(apk_path)
        if apk.is_valid_APK():
            apk_info.append(f)
            apk_info.append(get_file_md5(apk_path))
            apk_info.append(apk.get_app_name())
            apk_info.append(apk.get_package())
            apk_info.append(get_cert_md5(apk))
            apk_info.append(apk.get_androidversion_name())
    except Exception as e:
        print(f + ' ->>', e)

    return apk_info


def get_cert_md5(a):
    """
    获取证书md5
    :param a:
    :return:
    """
    cert_md5 = ''
    certs = set(a.get_certificates_der_v2() + [a.get_certificate_der(x) for x in a.get_signature_names()])
    for cert in certs:
        cert_md5 = hashlib.md5(cert).hexdigest()

    return cert_md5


class WriteData:
    def __init__(self, filename, navigation_bar):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.filename = filename + now
        self.font = Font(name='等线', size=11)
        self.fill = PatternFill("solid", fgColor="E0EEE0")
        self.title = navigation_bar

    def write_data(self, row=1, data=None, is_init=False):
        """
        写入数据
        :param row:
        :param data:
        :param is_init:
        :return:
        """
        if not is_init:
            for col in range(1, len(data) + 1):
                operate = self.ws.cell(row=row + 2, column=col, value=data[col - 1])
                operate.font = self.font
        else:
            for i in range(1, len(self.title) + 1):
                operate = self.ws.cell(row=row, column=i, value=self.title[i - 1])
                operate.font = self.font
                operate.fill = self.fill

        self.wb.save(os.path.join(download_dir, self.filename + '.xlsx'))


def rename2md5(path):
    start = time.time()
    if os.path.isdir(path):
        for root, dirs, files in os.walk(path):
            for i in range(len(files)):
                old_name = os.path.join(os.getcwd(), path, files[i])
                md5 = get_file_md5(old_name)
                new_name = os.path.join(os.getcwd(), path, md5 + '.apk')
                print('%s ->> %s' % (files[i] + '.apk', md5 + '.apk'))
                try:
                    os.rename(old_name, new_name)
                except FileExistsError as err:
                    print(err)
                    # 当有相同md5文件时执行去重操作
                    os.remove(old_name)
            print('\nTotal info:\n已对%d个文件进行重命名\t耗时%.2fs' % (len(files), time.time() - start))
    else:
        print('输入路径不是目录，请检查...\n%s' % path)


if __name__ == '__main__':
    main(sys.argv)
